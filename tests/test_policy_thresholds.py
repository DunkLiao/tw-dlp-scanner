import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from dlp_scanner import (
    DLPScannerApp,
    collect_text_detections,
    detect_or_rule_hits,
    scan_path,
    scan_path_with_or_hits,
    sort_result_rows,
)


def scan_text_content(content):
    with tempfile.TemporaryDirectory() as temp_dir:
        path = Path(temp_dir) / "sample.txt"
        path.write_text(content, encoding="utf-8")
        return scan_path(path)


def scan_text_content_with_or_hits(content):
    with tempfile.TemporaryDirectory() as temp_dir:
        path = Path(temp_dir) / "sample.txt"
        path.write_text(content, encoding="utf-8")
        return scan_path_with_or_hits(path)


def find_row(rows, hit_type):
    for row in rows:
        if row["hit_type"] == hit_type:
            return row
    return None


def find_row_prefix(rows, hit_type_prefix):
    for row in rows:
        if row["hit_type"].startswith(hit_type_prefix):
            return row
    return None


class PolicyThresholdTests(unittest.TestCase):
    def test_result_rows_sort_by_risk_high_medium_low_prompt(self):
        rows = [
            {"risk": "提示", "file_name": "prompt.txt"},
            {"risk": "中", "file_name": "medium.txt"},
            {"risk": "高", "file_name": "high.txt"},
            {"risk": "低", "file_name": "low.txt"},
        ]

        sorted_rows = sort_result_rows(rows)

        self.assertEqual(["高", "中", "低", "提示"], [row["risk"] for row in sorted_rows])
        self.assertEqual(["prompt.txt", "medium.txt", "high.txt", "low.txt"], [row["file_name"] for row in rows])

    def test_chinese_name_detection_keeps_clear_person_names(self):
        detected = collect_text_detections("王小明先生\n陳美華小姐\n李大華君")

        self.assertEqual(3, detected["4. 中文姓名"]["count"])
        self.assertEqual(["王小明", "陳美華", "李大華"], detected["4. 中文姓名"]["samples"])

    def test_chinese_name_detection_ignores_document_phrases(self):
        detected = collect_text_detections(
            "貸後監測作業規程：從被動審查轉向主動監控\n"
            "企業集團關係判定手冊：從股權到經營者\n"
            "一般流程主任說明"
        )

        self.assertNotIn("4. 中文姓名", detected)

    def test_or_rule_hits_report_any_raw_condition_below_policy_threshold(self):
        hits = detect_or_rule_hits("contact only_one@example.com")

        self.assertEqual(["10. 電子郵件"], hits)

    def test_scan_path_with_or_hits_reports_or_hits_without_policy_rows(self):
        rows, or_hits = scan_text_content_with_or_hits("contact only_one@example.com")

        self.assertEqual(["10. 電子郵件"], or_hits)
        self.assertEqual(1, len(rows))
        self.assertEqual(or_hits[0], rows[0]["hit_type"])
        self.assertEqual(1, rows[0]["hit_count"])

    def test_or_rule_hits_report_bank_account_without_keyword(self):
        hits = detect_or_rule_hits("account 123456789012")

        self.assertEqual(1, len(hits))
        self.assertTrue(hits[0].startswith("8. "))

    def test_scan_path_with_or_hits_reports_bank_account_without_policy_row(self):
        rows, or_hits = scan_text_content_with_or_hits("account 123456789012")

        self.assertEqual(1, len(or_hits))
        self.assertTrue(or_hits[0].startswith("8. "))
        self.assertEqual(1, len(rows))
        self.assertEqual(or_hits[0], rows[0]["hit_type"])
        self.assertEqual(1, rows[0]["hit_count"])
        self.assertEqual("123456789012", rows[0]["sample"])
        self.assertNotIn("*", rows[0]["sample"])

    def test_policy_rows_keep_full_email_sample(self):
        content = "\n".join(f"user{index:02d}@example.com" for index in range(40))

        row = find_row(scan_text_content(content), "10. 電子郵件")

        self.assertIsNotNone(row)
        self.assertIn("user00@example.com", row["sample"])
        self.assertNotIn("***", row["sample"])

    def test_excel_report_keeps_full_sample(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            report_path = Path(temp_dir) / "report.xlsx"
            fake_app = type("FakeApp", (), {})()
            fake_app.export_rows = [{
                "掃描時間": "2026-07-07 10:00:00",
                "風險等級": "提示",
                "處置": "額外OR提示",
                "檔案名稱": "aa.txt",
                "命中類型": "8. 銀行存款帳號與關鍵字",
                "命中次數": 1,
                "命中範例": "123456789012",
                "完整路徑": r"D:\Download\aa.txt",
            }]

            DLPScannerApp.write_excel_report(fake_app, report_path)
            workbook = load_workbook(report_path)
            sheet = workbook["掃描結果"]

            self.assertEqual("123456789012", sheet.cell(row=2, column=7).value)

    def test_excel_report_sorts_rows_by_risk(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            report_path = Path(temp_dir) / "report.xlsx"
            fake_app = type("FakeApp", (), {})()
            fake_app.export_rows = [
                {"掃描時間": "2026-07-07 10:00:00", "風險等級": "提示", "處置": "額外OR提示", "檔案名稱": "d.txt"},
                {"掃描時間": "2026-07-07 10:00:00", "風險等級": "低", "處置": "僅記錄", "檔案名稱": "c.txt"},
                {"掃描時間": "2026-07-07 10:00:00", "風險等級": "高", "處置": "記錄並阻擋", "檔案名稱": "a.txt"},
                {"掃描時間": "2026-07-07 10:00:00", "風險等級": "中", "處置": "僅記錄", "檔案名稱": "b.txt"},
            ]

            DLPScannerApp.write_excel_report(fake_app, report_path)
            workbook = load_workbook(report_path)
            sheet = workbook["掃描結果"]

            self.assertEqual(["高", "中", "低", "提示"], [sheet.cell(row=index, column=2).value for index in range(2, 6)])

    def test_or_rule_hits_report_bank_account_condition(self):
        hits = detect_or_rule_hits("客戶往來明細查詢 123456789012")

        self.assertEqual(["8. 銀行存款帳號與關鍵字"], hits)

    def test_email_thresholds_use_actual_count_for_low_medium_high(self):
        for count, expected_risk, expected_action in [
            (40, "低", "僅記錄"),
            (60, "中", "僅記錄"),
            (80, "高", "記錄並阻擋"),
        ]:
            content = "\n".join(f"user{index}@example.com" for index in range(count))

            row = find_row(scan_text_content(content), "10. 電子郵件")

            self.assertIsNotNone(row)
            self.assertEqual(expected_risk, row["risk"])
            self.assertEqual(expected_action, row["action"])
            self.assertEqual(count, row["hit_count"])

    def test_identity_number_thresholds_are_medium_at_one_and_high_at_two(self):
        one_id_row = find_row(scan_text_content("A123456789"), "5. 台灣身分證字號")
        two_id_row = find_row(scan_text_content("A123456789\nB223456789"), "5. 台灣身分證字號")

        self.assertEqual("中", one_id_row["risk"])
        self.assertEqual("僅記錄", one_id_row["action"])
        self.assertEqual(1, one_id_row["hit_count"])
        self.assertEqual("高", two_id_row["risk"])
        self.assertEqual("記錄並阻擋", two_id_row["action"])
        self.assertEqual(2, two_id_row["hit_count"])

    def test_bank_account_policy_requires_keyword_and_12_to_14_digits(self):
        for account in ["123456789012", "1234567890123", "12345678901234"]:
            row = find_row(scan_text_content(f"客戶往來明細查詢 {account}"), "8. 銀行存款帳號與關鍵字")

            self.assertIsNotNone(row)
            self.assertEqual("高", row["risk"])
            self.assertEqual("記錄並阻擋", row["action"])
            self.assertEqual(1, row["hit_count"])

        self.assertIsNone(find_row(scan_text_content("客戶往來明細查詢 12345678901"), "8. 銀行存款帳號與關鍵字"))
        self.assertIsNone(find_row(scan_text_content("客戶往來明細查詢 123456789012345"), "8. 銀行存款帳號與關鍵字"))
        self.assertIsNone(find_row(scan_text_content("帳號 123456789012"), "8. 銀行存款帳號與關鍵字"))

    def test_combined_personal_data_thresholds_follow_policy(self):
        low_content = "\n".join([
            "王小明先生 0912-345-678",
            "陳美華小姐 0988-765-432",
        ])
        medium_content = low_content + "\n李大華先生"
        high_content = "\n".join([
            "王小明先生 0912-345-678",
            "陳美華小姐 0988-765-432",
            "李大華先生",
            "林志強先生",
            "張雅婷小姐",
        ])

        low_row = find_row(scan_text_content(low_content), "1. 組合式個資")
        medium_row = find_row(scan_text_content(medium_content), "1. 組合式個資")
        high_row = find_row(scan_text_content(high_content), "1. 組合式個資")

        self.assertEqual("低", low_row["risk"])
        self.assertEqual("中", medium_row["risk"])
        self.assertEqual("高", high_row["risk"])
        self.assertEqual("記錄並阻擋", high_row["action"])


if __name__ == "__main__":
    unittest.main()
